"""
컬럼 설계서 자동 생성
채팅에서 수집한 컬럼 정보 → Excel/CSV 다운로드
"""
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import io, csv, json
from ..core.database import db_cursor

router = APIRouter()

class ColumnSpec(BaseModel):
    column_name: str          # 컬럼명 (영문약어)
    korean_name: str          # 한글명
    data_type: Optional[str] = ""
    length: Optional[str]    = ""
    decimal: Optional[str]   = ""
    storage_format: Optional[str] = ""
    display_format: Optional[str] = ""
    description: Optional[str] = ""
    nullable: Optional[str]  = "N"
    primary_key: Optional[bool] = False

class ExportRequest(BaseModel):
    columns: List[ColumnSpec]
    table_name: Optional[str] = "TABLE_NAME"
    format: Optional[str] = "xlsx"   # xlsx | csv


@router.post("/export/columns")
async def export_columns(req: ExportRequest):
    if req.format == "csv":
        return _export_csv(req)
    return _export_xlsx(req)


def _export_csv(req: ExportRequest):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["컬럼명(영문)", "한글명", "데이터타입", "길이", "소수점",
                     "저장형식", "표현형식", "설명", "NOT NULL", "PK"])
    for col in req.columns:
        writer.writerow([
            col.column_name, col.korean_name, col.data_type, col.length,
            col.decimal, col.storage_format, col.display_format,
            col.description, col.nullable, "Y" if col.primary_key else "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{req.table_name}_columns.csv"'},
    )


def _export_xlsx(req: ExportRequest):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = req.table_name[:31]

    # 스타일
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2563EB")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = ["컬럼명(영문)", "한글명", "데이터타입", "길이", "소수점",
               "저장형식", "표현형식", "설명", "NOT NULL", "PK"]
    WIDTHS  = [20, 20, 14, 8, 8, 25, 25, 40, 10, 6]

    # 제목
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"{req.table_name} 컬럼 정의서"
    title_cell.font  = Font(bold=True, size=13)
    title_cell.alignment = center
    title_cell.fill = PatternFill("solid", fgColor="EFF6FF")
    ws.row_dimensions[1].height = 28

    # 헤더
    for col_idx, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 20

    # 데이터
    alt_fill = PatternFill("solid", fgColor="F8FAFF")
    for row_idx, col in enumerate(req.columns, 3):
        row_data = [
            col.column_name, col.korean_name, col.data_type, col.length,
            col.decimal, col.storage_format, col.display_format,
            col.description, col.nullable, "Y" if col.primary_key else "",
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or "")
            cell.alignment = center if col_idx in [3,4,5,9,10] else left
            cell.border    = border
            if fill: cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    # 틀 고정
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{req.table_name}_columns.xlsx"'},
    )


@router.get("/export/columns/template")
def get_template():
    """공통표준 기반 컬럼 정보 조회 (여러 용어 한꺼번에)"""
    pass


@router.post("/export/columns/lookup")
async def lookup_columns(body: dict):
    """용어명 목록을 받아 공통표준에서 컬럼 정보를 일괄 조회"""
    term_names = body.get("terms", [])
    if not term_names:
        return []

    results = []
    with db_cursor() as cur:
        for name in term_names:
            cur.execute("""
                SELECT t.term_name, t.term_abbr, t.term_desc, t.domain_name,
                       d.data_type, d.data_length, d.data_decimal,
                       d.storage_format, d.display_format
                FROM std_term t
                LEFT JOIN std_domain d ON d.domain_name = t.domain_name
                WHERE t.term_name = %s
                LIMIT 1
            """, (name,))
            row = cur.fetchone()
            if row:
                r = dict(row)
                results.append(ColumnSpec(
                    column_name   = r["term_abbr"] or "",
                    korean_name   = r["term_name"],
                    data_type     = r["data_type"] or "",
                    length        = str(r["data_length"]) if r["data_length"] else "",
                    decimal       = str(r["data_decimal"]) if r["data_decimal"] else "",
                    storage_format= r["storage_format"] or "",
                    display_format= r["display_format"] or "",
                    description   = r["term_desc"] or "",
                ))
            else:
                results.append(ColumnSpec(column_name="", korean_name=name))
    return results
